from ctypes import wstring_at
import openpyxl

#1) 엑셀 만들기
wb = openpyxl.Workbook()

#2) 엑셀 워크시트 만들기
ws = wb.create_sheet('오징어게임') #('워크시트 타이틀')

#3) 데이터 추가하기
ws
ws['B1'] = '성명'

ws['A2'] = '1'
ws['B2'] = '오일남'

#4) 엑셀 저장하기
wb.save(r'D:\factory\ws_sts_4151\seoul\public\crawling2\02_파이썬엑셀다루기\참가자_data.xlsx') # \ 는 이스케이프 문자를 만드는 역할이라 그냥 문자 형태로 인식하기 위해 문자열 기호 r을 붙인다
                                                                  # 경로를 전부 \\ 두개를 or 전부 /로 통일시킨다