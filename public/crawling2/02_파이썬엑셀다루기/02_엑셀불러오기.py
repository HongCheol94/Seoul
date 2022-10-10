import openpyxl

fpath = r'/Users/kims/Desktop/java1/java/factory/git/seoul/public/crawling2/02_파이썬엑셀다루기\참가자_data.xlsx'

# 1 ) 엑셀 불러오기
wb = openpyxl.load_workbook(fpath) #컴퓨터는 반복을 안좋아해서 주소를 fpath라는 변수에 넣는다.
# 2 ) 엑셀 시트선택
ws = wb['오징어게임']

# 3 ) 데이터 수집하기
ws['A3'] = 456
ws['B3'] = '성기훈'

# 4 ) 엑셀 저장하기
wb.save(fpath)